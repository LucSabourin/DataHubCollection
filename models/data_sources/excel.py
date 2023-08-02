from os import remove
from posixpath import split
from openpyxl.descriptors.base import DateTime
import pandas as pd
import json
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from datetime import datetime

from models.data_sources.datasource_classes import FileSource, DataSource
import models.scans.scan_excel as scan
from models.scans.scan_generic import duplicate_headers
from models.exceptions.exceptions import ServerError, UserError


class ExcelSource(FileSource):
    """
    """

    message_templates = {
            'sheet_not_in_book': ('Sheet {1} not in source {0}.', 4),
            'formula_error': ('Formula {1} in cell {3} of source {0} resulted in error {2}.', 3),
            'formula': ('Formula {1} in source {0}.', 0),
            'merged_cells': ('Cells in range {1} found in source {0} have been unmerged, so contents have been stored in cell {2} and all other cells in the range are empty.', 0),
            'hidden_rows': ('Hidden rows {1} in source {0} have been unhidden, so their contents have been included.', 2),
            'hidden_columns': ('Hidden column groups starting at {1} in source {0} have been unhidden, so their contents have been included.', 2),
            'no_sheets_in_book': ('None of the supplied sheets are in source {0}', 88),
            'could_not_load_into_openpyxl': ('Could not load source {0} as workbook (xls, xlsx, xlsm).', 99)
        }

    resource_type = 'excel'

    version = 1.0

    def __init__(self,
        catalogue_key: str,
        location: str = '',
        connections: list = None,
        index_fields: dict = None,
        remove_empty: bool = False,
        read_only: bool = False,
        data_fresh_date: datetime = None,
        ingested_location: str = '',
        ingested_at: str = '',
        ingested: bool = False,
        tables: list = None,
        resource_type: str = '',
        version: float = 1.0,
        metadata: dict = None,
        file_name: str = '',
        location_type: str = '',
        tags: dict = None
    ) -> None:
        FileSource.__init__(self,
            catalogue_key=catalogue_key,
            location=location,connections=connections,
            index_fields=index_fields,
            remove_empty=remove_empty,
            read_only=read_only,
            data_fresh_date=data_fresh_date,
            ingested_location=ingested_location,
            ingested_at=ingested_at,
            ingested=ingested,
            resource_type=resource_type,
            version=version,
            metadata=metadata,
            file_name=file_name,
            location_type=location_type,
            tags=tags
        )
        if tables is not None:
            self.tables = tables
        else:
            self.tables = []
        
    def scan(self, messages: list):
        """Builds metadata for the file which governs what processes
        be done to get the data into a provisionable asset.
        """
        try:
            book_data = scan.open_wkbk(self.file, True)
        except:
            msg = self.message('could_not_load_into_openpyxl', [self.file_name], self.message_templates)
            messages.append(msg)
            messages.append(self.message_severity)
            raise ServerError(msg)
        book_sheets = book_data.sheetnames
        if len(self.tables) > 0:
            missing_sheets = [sheet for sheet in self.tables if sheet not in book_sheets]
            if len(missing_sheets) > 0:
                # Generate error message for each sheet in filter list
                # not in workbook.
                for sheet_name in missing_sheets:
                    msg = self.message('sheet_not_in_book', [self.file_name, sheet_name], self.message_templates)
                    messages.append(msg)
                    self.tables.remove(sheet_name)

                if len(self.tables) == 0:
                    msg = self.message('no_sheets_in_book', [self.file_name], self.message_templates)
                    messages.append(msg)
                    messages.append(self.message_severity)
                    raise UserError(msg)
                    
                self.metadata = {sheet: {} for sheet in self.tables}
            else:
                # If filter list was not empty and all sheets were in
                # the workbook, filter worksheets.
                self.metadata = {sheet: {} for sheet in book_sheets if sheet in self.tables}
        else:
            # If no sheets in filter list, use all sheets in workbook.
            self.tables = [sheet for sheet in book_sheets]
            self.metadata= {sheet: {} for sheet in book_sheets}

        book_formula = scan.open_wkbk(self.file, False)
        update_file = False
        for sheet_name in self.tables:
            sheet = book_data[sheet_name]
            # Min/Max Row and Column Indices
            self.metadata[sheet_name]['first_row'] = scan.sheet_min_row(sheet)
            self.metadata[sheet_name]['last_row'] = scan.sheet_max_row(sheet)
            self.metadata[sheet_name]['first_column'] = scan.sheet_min_col(sheet)
            self.metadata[sheet_name]['last_column'] = scan.sheet_max_col(sheet)
            # Number of empty rows to the top of data
            if self.metadata[sheet_name]['first_row'] > 1:
                msg = self.message('top_rows_empty', [sheet_name, self.metadata[sheet_name]['first_row'] - 1], self.message_templates)
                messages.append(msg)
            # Number of empty columns to the left of data
            if self.metadata[sheet_name]['first_column'] > 1:
                msg = self.message('left_columns_empty', [sheet_name, self.metadata[sheet_name]['first_column'] - 1], self.message_templates)
                messages.append(msg)
            # Merged Cells
            merged_cells = scan.sheet_merged_cells(sheet)
            if merged_cells is not None:
                for cell_range, info in merged_cells.items():
                    msg = self.message('merged_cells', [sheet_name, cell_range, info[0]], self.message_templates)
                    messages.append(msg)
            # Header info
            field_header_row = scan.sheet_field_headers_row(
                sheet=sheet, 
                min_row=self.metadata[sheet_name]['first_row'],
                max_row=self.metadata[sheet_name]['last_row'],
                min_col=self.metadata[sheet_name]['first_column'],
                max_col=self.metadata[sheet_name]['last_column']
            )
            # Has Headers in file
            if field_header_row != -1:
                # Base Header Info.
                self.metadata[sheet_name]['field_header_row'] = field_header_row
                headers = scan.sheet_field_headers(
                    sheet=sheet,
                    header_row=field_header_row,
                    min_col=self.metadata[sheet_name]['first_column'],
                    max_col=self.metadata[sheet_name]['last_column']
                )
                # Field Formats
                self.metadata[sheet_name]['field_formats'] = scan.sheet_field_formats(
                    sheet=sheet,
                    min_row_data=field_header_row + 1,
                    min_col=self.metadata[sheet_name]['first_column'],
                    max_col=self.metadata[sheet_name]['last_column']
                )
                # Duplicate Field Headers
                dupe_headers = duplicate_headers(
                    headers=headers,
                    min_col = self.metadata[sheet_name]['first_column']
                )
                if dupe_headers is not None:
                    for header, positions in dupe_headers:
                        msg = self.message('duplicate_headers', [sheet_name, header, len(positions), positions], self.message_templates)
                        messages.append(msg)
                # Hierarchical Field Headers
                hierarchical_headers = scan.sheet_hierarchical_fields(
                    sheet=sheet,
                    min_row=self.metadata[sheet_name]['first_row'],
                    header_row=field_header_row,
                    min_col=self.metadata[sheet_name]['first_column'],
                    max_col=self.metadata[sheet_name]['last_column']
                )
                if len([header for header in hierarchical_headers if '.' in str(header or '')]) > 0:
                    headers = hierarchical_headers
                    msg = self.message('hierarchical_headers', [sheet_name, hierarchical_headers], self.message_templates)
                    messages.append(msg)
                # Missing Field Headers
                if None in headers:
                    missing_headers = [num + self.metadata[sheet_name]['first_column'] for num, header in enumerate(headers) if header is None]
                    msg = self.message('missing_headers', [sheet_name, [get_column_letter(pos) for pos in missing_headers]], self.message_templates)
                    messages.append(msg)
                self.metadata[sheet_name]['headers'] = headers
            # Does Not Have Headers in file
            else:
                # Add empty row above data so it does not get cut off
                if self.metadata[sheet_name]['first_row'] == 1:
                    book_data[sheet_name].insert_rows(1)
                    update_file = True
                # Base Header Info
                headers = []
                for num in range(self.metadata[sheet_name]['first_column'], self.metadata[sheet_name]['last_column'] + 1):
                    headers.append('Unnamed: ' + str(num))
                self.metadata[sheet_name]['headers'] = headers
                
                # Field formats
                self.metadata[sheet_name]['field_formats'] = scan.sheet_field_formats(
                    sheet=sheet,
                    min_row_data=self.metadata[sheet_name]['first_row'],
                    min_col=self.metadata[sheet_name]['first_column'],
                    max_col=self.metadata[sheet_name]['last_column']
                )
                # Missing Field Headers
                missing_headers = [num + self.metadata[sheet_name]['first_column'] for num in range(0, len(headers))]
                msg = self.message('missing_headers', [sheet_name, [get_column_letter(pos) for pos in missing_headers]], self.message_templates)
                messages.append(msg)
            # Hidden Columns
            hidden_columns = scan.sheet_hidden_columns(
                sheet=sheet,
                min_col=self.metadata[sheet_name]['first_column'],
                max_col=self.metadata[sheet_name]['last_column']
            )
            if len(hidden_columns) > 0:
                msg = self.message('hidden_columns', [sheet_name, hidden_columns], self.message_templates)
                messages.append(msg)
            # Hidden Rows
            hidden_rows = scan.sheet_hidden_rows(
                sheet=sheet,
                min_row=self.metadata[sheet_name]['first_row'],
                max_row=self.metadata[sheet_name]['last_row']
            )
            if len(hidden_rows) > 0:
                msg = self.message('hidden_rows', [sheet_name, hidden_rows], self.message_templates)
                messages.append(msg)
            # Formulas
            sheet_formula = book_formula[sheet_name]
            formulas = scan.sheet_formulas(
                sheet=sheet_formula,
                min_row_data=self.metadata[sheet_name]['first_row'],
                max_row=self.metadata[sheet_name]['last_row'],
                min_col=self.metadata[sheet_name]['first_column'],
                max_col=self.metadata[sheet_name]['last_column']
            )
            if formulas is not None:
                # Formula Errors
                for cell_ref, formula in formulas.items():
                    if scan.sheet_check_formula_error(
                        sheet=sheet,
                        row_index=formula[0],
                        col_index=formula[1]
                    ) is True:
                        msg = self.message('formula_error', [sheet_name, formula[2], sheet.cell(formula[0], formula[1]).value, cell_ref], self.message_templates)
                        messages.append(msg)

        # Saves any inserted rows for sheets missing headers with data in 
        # first row to capture first row of data.
        if update_file is True:
            self.file = BytesIO(save_virtual_workbook(book_data))
        # Jsonifies metadata.
        self.metadata = json.loads(json.dumps(self.metadata))

    def to_dataframe(self, messages: list) -> None:
        """Converts data from ByteIO file-like object to a pandas
        dataframe.
        """
        for sheet_name in self.tables:
            # Position of metadata for this sheet
            # Import this sheet from file
            schema = {'io': self.file, 'sheet_name': sheet_name}
            
            schema['names'] = self.metadata[sheet_name]['headers']
            
            # Skip rows above data (hierarchical fields and field
            # headers) included in if/else statement above
            if 'field_header_row' in self.metadata[sheet_name].keys():
                schema['skiprows'] = range(0, self.metadata[sheet_name]['field_header_row'] - 1)
            elif self.metadata[sheet_name]['first_row'] > 1:
                schema['skiprows'] = range(0, self.metadata[sheet_name]['first_row'] - 1)

            # Use columns containing data
            schema['usecols'] = range(
                self.metadata[sheet_name]['first_column'] - 1,
                self.metadata[sheet_name]['last_column']
            )

            try:
                df = pd.read_excel(**schema)
            except:
                msg = self.message('could_not_load_into_pandas', [self.file_name], self.message_templates)
                messages.append(msg)
                messages.append(self.message_severity)
                raise ServerError(msg)

            self.data[sheet_name] = df

    def clean_up(self, messages: list) -> list:
        """Run any necessary clean up, save logs for warnings if
        anything structurally changes or the values of data changes as
        a result.
        Ex. removing Null rows or columns.
        """
        messages = []
        for sheet_name, df in self.data.items():
            # Columns
            for column_index, column in enumerate(df.columns):
                if df[df.columns[column_index]].count() == 0:
                    msg = self.message('empty_column', [sheet_name, column], self.message_templates)
                    messages.append(msg)
                    if self.remove_empty is True:
                        self.data[sheet_name] = df.drop(column, axis=1)
            # Rows
            for row_index, row in df.iterrows():
                if row.isnull().all():
                    msg = self.message('empty_row', [sheet_name, row_index], self.message_templates)
                    messages.append(msg)
            if self.remove_empty is True:
                self.data[sheet_name] = df.dropna(how='all')
        if len(messages) > 0:
            return messages

    def message(self, message_template: str, params: list, message_templates: dict) -> str:
        """
        """
        
        if message_template in message_templates.keys():
            if message_templates[message_template][1] > self.message_severity:
                self.message_severity = message_templates[message_template][1]
            return str(message_templates[message_template][1]) + ' -> ' + message_templates[message_template][0].format(*params)
        else:
            return super().message(message_template, params, super().message_templates)

    def store_oc_data(self, messages: list) -> None:
        first_version = bool(len(self.ingested_location) == 0)

        folder = self.assign_folder(messages=messages)
        self.ingested_location = folder
        if first_version is False:
            old_ingested = json.load(self.get_file(location=self.ingested_location, messages=messages))
            old_ingested['location'] = self.ingested_location
        else:
            old_ingested = {'ingested_locations': []}
            old_ingested['location'] = ''

        ingested = {}
        ingested['ingested_locations'] = []

        ### TESTING FAILED CLEANUP
        #break_test = 0

        for sheet_name, df in self.data.items():
            data = json.dumps({sheet_name: json.loads(df.to_json(orient='split'))})

            file_name = folder + '/' + sheet_name + '.json'

            tags = {}
            tags['sheet_name'] = sheet_name
            tags['catalogue_key'] = self.catalogue_key

            old_url = [url for url in old_ingested['ingested_locations'] if sheet_name + '.json' in url]
            if len(old_url) == 1:
                old_url = old_url[0]
            else:
                old_url = ''

            url = self.post_file(file_name=file_name, data=data, messages=messages, tags=tags, old_url=old_url)
            ingested['ingested_locations'].append(url)

            if len(old_url) > 0:
                old_ingested['ingested_locations'].remove(old_url)

            ### TESTING FAILED CLEANUP
            #if break_test > 0:
            #    messages.append('99 -> Break Test Error')
            #    messages.append(99)
            #    raise ServerError()
            #break_test += 1

        for location in old_ingested['ingested_locations']:
            self.delete_file(location=location, messages=messages)

        self.ingested = True
        self.ingested_at = datetime.now().strftime('%Y/%m/%d %H:%M:%S')
        ingested['ingested'] = self.ingested
        ingested['ingested_at'] = self.ingested_at

        self.ingested_location = self.post_file(
            file_name=folder + '/:ingested.json',
            data=json.dumps(ingested),
            messages=messages,
            tags={'catalogue_key': self.catalogue_key},
            old_url=old_ingested['location']
        )

        self.updating = False

    def failed_clean_up(self, messages: list) -> None:
        for sheet in self.tables:
            location = self.ingested_location + '/' + sheet + '.json'
            self.delete_file(location=location, messages=messages)
        self.ingested_location = ''
        
    def retrieve_data(self, tables: list, messages: list) -> None:
        if len(tables) > 0:
            missing_sheets = [sheet for sheet in tables if sheet not in self.tables]
            if len(missing_sheets) > 0:
                for sheet_name in missing_sheets:
                    msg = self.message('sheet_not_in_book', [self.file_name, sheet_name], self.message_templates)
                    messages.append(msg)
                    tables.remove(sheet_name)
            if len(tables) == 0:
                msg = self.message('no_sheets_in_book', [self.file_name], self.message_templates)
                messages.append(msg)
                messages.append(self.message_severity)
                raise UserError(msg)
        else:
            tables = self.tables

        ingested = json.load(self.get_file(location=self.ingested_location, messages=messages))

        for sheet_name in tables:
            if sheet_name not in self.data.keys():
                url = [url for url in ingested['ingested_locations'] if sheet_name in url][0]
                data = json.load(self.get_file(location=url, messages=messages))[sheet_name]

                df = pd.DataFrame(data=data['data'], columns=data['columns'], index=data['index'])
                self.data[sheet_name] = df

    def get_response(self, messages: list, tables: list = None, fields: list = None, table_field: dict = None, field_value: dict = None, orient = 'records'):
        """
        """
        empty_tables = False
        if tables is None or tables == []:
            tables = []
            empty_tables = True
        if fields is None or fields == []:
            fields = []
        if table_field is None or table_field == []:
            table_field = {}
        if field_value is None or field_value == []:
            field_value = {}

        data = {}
        self.retrieve_data(tables=tables, messages=messages)
        
        missing_fields = {}
        for field in fields:
            missing_fields[field] = 0

        empty_fields = True
        if len(fields) > 0:
            empty_fields = False
        for sheet_name, df in self.data.items():
            if (sheet_name in tables) or (empty_tables):
                data[sheet_name] = json.loads(df.to_json(orient=orient))

        messages.append(self.message_severity)
        return data

    def delete_response(self, messages: list):
        ingested = json.load(self.get_file(location=self.ingested_location, messages=messages))
        for location in ingested['ingested_locations']:
            self.delete_file(location=location, messages=messages)

        self.delete_file(location=self.ingested_location, messages=messages)

    def serialize(self):
        """*** Yet to be tested***
        """
        data_source = {}
        data_source['location'] = self.location
        data_source['tables'] = self.tables
        data_source['index_fields'] = self.index_fields
        data_source['metadata'] = self.metadata
        data_source['ingested_location'] = self.ingested_location
        data_source['ingested_at'] = self.ingested_at
        data_source['ingested'] = self.ingested
        data_source['resource_type'] = self.resource_type
        data_source['version'] = self.version
        data_source['file_name'] = self.file_name
        data_source['catalogue_key'] = self.catalogue_key
        data_source['location_type'] = self.location_type
        data_source['tags'] = self.tags
        return data_source
