import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet 
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from io import BytesIO

def open_wkbk(file: BytesIO, formula_values: bool = True) -> Workbook:
    """Opens an excel workbook and returns the openpyxl Excel Workbook
    object.

    Parameters:
    -----------
    file : BytesIO
        BytesIO file like object representing an excel file, either
        from cloud source or local network.

    formula_values : bool
        Whether the values of any worksheet functions should be
        calculated; if True includes values, if False includes
        formulas in lieu of values.

    Returns:
    --------
    Workbook
        From openpyxl.workbook.workbook.Workbook package after opening
        the BytesIO file like object using openpyxl.
    """
    workbook = openpyxl.load_workbook(file, data_only=formula_values)
    
    return workbook

def check_sheets_in_wkbk(wkbk: Workbook, sheets: list = None) -> list:
    """Checks that the worksheets in the list of sheets are in the
    workbook. If the supplied list of sheets is empty, uses all sheets
    in the workbook.

    Parameters
    ----------
    wkbk : Workbook
        A workbook already instanced by openpyxl (open_wkbk function).

    sheets : list
        A list of sheets to include when filtering. If empty, no
        filtering takes place.

    Returns
    -------
    list -> [WorkSheet]
        Returns a list of Worksheet objects.
    """
    if sheets is None:
        sheets = []
    filter_sheets = False
    if len(sheets) > 0:
        filter_sheets = True

    use_sheets = []
    for sheet in wkbk.sheetnames:
        if filter_sheets is False or sheet in sheets:
            use_sheets.append(wkbk[sheet])

    return use_sheets

def sheet_min_row(sheet: Worksheet, limit_search: int = 100) -> int:
    """Checks for the first row containing a non-empty cell in the
    sheet and returns the row index (1-based) ofthe first row
    containing data.

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to find first row containing data/non empty cells.

    limit_search : int (Optional)
        Limits the search to the first n rows, default is 100.

    Returns
    -------
    int
        1-based row index containing data. If no row is found, or the
        search limit is reached, returns -1.
    """
    for row_index, row in enumerate(sheet.rows):

        # If the length of a list containing cell values with length
        # greater than 0 is itself greater than 0, return the row_index
        #  + 1.
        if len([cell.value for cell in row if len(str(cell.value or '')) > 0]) > 0:
            return row_index + 1
        elif row_index >= limit_search:
            break
    
    return -1

def sheet_min_col(sheet: Worksheet, limit_search: int = 100) -> int:
    """Checks for the first column containing a non-empty cell in the
    sheet and returns the column index (1-based) of the first column
    containing data.

    If no column can be found containing data, returns -1.

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to find first column containing data/non empty cells.

    limit_search : int (Optional)
        Limits the search to the first n columns, default is 100.

    Returns
    -------
    int
        1-based column index for the first column containing data. If
        no column is found, or the search limit is reached, returns -1.
    """
    for column_index, column in enumerate(sheet.columns):

        # If the length of a list containing cell values with length
        # greater than 0 is itself greater than 0, return the 
        # column_index + 1.
        if len([cell.value for cell in column if len(str(cell.value or '')) > 0]) > 0:
            return column_index + 1
        elif column_index >= limit_search:
            break

    return -1

def sheet_max_row(sheet: Worksheet) -> int:
    """Returns the row index (1-based) for the last row containing data
    in the worksheet supplied.

    Uses openpyxl.worksheet.worksheet.Worksheet.max_row, which has
    reportedly caused some issues when contents in a row are cleared
    but the rows have not been deleted.

    This issue is caused by Excel and the way it stores information.

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to find last row containing data/non empty cells.

    Returns
    -------
    int
        1-based index for the last row containing data.
    """
    return sheet.max_row

def sheet_max_col(sheet: Worksheet) -> int:
    """Returns the column index (1-based) for the last column 
    containing data in the worksheet supplied.

    Uses openpyxl.worksheet.worksheet.Worksheet.max_column, which has
    reportedly caused some issues when contents in a row are cleared
    but the rows have not been deleted.

    This issue is caused by Excel and the way it stores information.

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to find last row containing data/non empty cells.

    Returns
    -------
    int
        1-based index for the last column containing data.
    """
    return sheet.max_column

def sheet_field_headers_row(sheet: Worksheet, min_col: int, max_col: int, min_row: int, max_row: int) -> int:
    """Returns the row containing field headers in the sheet.

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to find for containing field headers.
    
    min_col : int
        1-based index of first column containing data.

    max_col : int
        1-based index of last column containing data.

    min_row : int
        1-based index of first row containing data.

    max_row : int
        1-based index of last row containing data.

    Returns
    -------
    int
        1-based index of row containing bottom most field headers,
        where there are hierarchical fields, otherwise the row
        containing the field headers.
    """
    for column_index in range(min_col, max_col + 1):
        columns_raw = sheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=min_row,
            max_row=max_row
        )
        column = [cell.value for cell in [cell for cell in columns_raw][0]]
        header_type = type(None)
        header_row = 0
        for row_index, cell in enumerate(column):
            if type(cell) == type('') and header_type == type(None):
                header_type = type('')
            elif header_type == type('') and row_index <= 9:
                if type(cell) != header_type:
                    header_row = row_index + min_row - 1
                    break
                elif cell[0] in ['=', '#']:
                    header_row = row_index + min_row - 1
                    break
                elif '\n' in cell:
                    header_row = row_index + min_row - 1
                    break
            else:
                break
                
        if len([cell for cell in column if type(cell) != type('')]) < len(column) and header_row > 0:
            return header_row
    
    return -1

def sheet_field_headers(sheet: Worksheet, header_row: int, min_col: int, max_col: int) -> list:
    """Returns the field headers based on the field header row in the
    sheet.

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to find for containing field headers.

    header_row : int
        1-based index of row containing bottom most field header (for
        hierarchical field headers)/field headers full stop.

    min_col : int
        1-based index of first column containing data.

    max_col : int
        1-based index of last column containing data.

    Returns
    -------
    list -> [<field_header: str>]
        Returns a list of field headers (strings).
    """
    headers_raw = sheet.iter_rows(
        min_row=header_row,
        max_row=header_row,
        min_col=min_col,
        max_col=max_col
    )

    return [cell.value for cell in [cell for cell in headers_raw][0]]

def sheet_field_formats(sheet: Worksheet, min_row_data: int, min_col: int, max_col: int) -> list:
    """Selects the first row of data and returns the cell formatting of
    each cell in a list.

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to find for field cell format.

    min_row_data : int
        1-based index of first row containing data (not field headers).
    
    min_col : int
        1-based index of first column containing data.

    max_col : int
        1-based index of last column containing data.

    Returns
    -------
    list -> [<cell_format: str>]
        Returns a list of cell formats (strings).
    """
    formats_raw = sheet.iter_rows(
        min_row=min_row_data,
        max_row=min_row_data,
        min_col=min_col,
        max_col=max_col
    )
    return [cell.style for cell in [cell for cell in formats_raw][0]]

def sheet_hidden_columns(sheet: Worksheet, min_col: int, max_col: int) -> list:
    """Finds first hidden column in a group of hidden columns and
    returns a list of 1-based column indices for each.

    Note that due to the way that Excel stores the information, there
    isn't a way to get the information for all other columns hidden in
    a group.

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to find hidden columns.

    min_col : int
        1-based index of first column containing data.

    max_col : int
        1-based index of last column containing data.

    Returns
    -------
    list -> [<col_index: int>]
        Returns a list of 1-based column indices refering to first
        hidden column in a group of hidden columns.
    """
    hidden = []
    for col_index in range(min_col, max_col + 1):
        col = get_column_letter(col_index)
        if sheet.column_dimensions[col].hidden is True:
            hidden.append(col_index)
    return hidden

def sheet_hidden_rows(sheet: Worksheet, min_row: int, max_row: int) -> list:
    """Finds hidden rows and returns a list of 1-based column indices
    for each.

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to find hidden rows.

    min_row : int
        1-based index of first row containing data.

    max_row : int
        1-based index of last row containing data.

    Returns
    -------
    list -> [<row_index: int>]
        Returns a list of 1-based column indices refering to all hidden
        rows.
    """
    hidden = []
    for row_index in range(min_row, max_row + 1):
        if sheet.row_dimensions[row_index].hidden is True:
            hidden.append(row_index)
    return hidden

def sheet_merged_cells(sheet: Worksheet) -> dict:
    """Returns a dictionary of range: 1-based index coordinates of
    start/stop locations for merged cells in a worksheet.

    If no merged cells in a sheet, returns None.

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to find merged cells.

    Returns
    -------
    dict -> {<merged_range: str>: [(<row_start: int>, <col_start: int>), (<row_end: int>, <col_end: int>)]}
        Returns a dictionary containing ranges ('A1:B2' for example) as
        keys and a list of tuples referring to the start (top left) and
        end (bottom right) of a merged cell.
    """
    merged = [str(merged_range or '') for merged_range in sheet.merged_cells.ranges]
    merged_cells = {}
    if len(merged) > 0:
        for range in merged:
            start = coordinate_to_tuple(range.split(':')[0])
            end = coordinate_to_tuple(range.split(':')[1])
            merged_cells[range] = [range.split(':')[0], start, end]
        return merged_cells
    return None

def sheet_formulas(sheet: Worksheet, min_row_data: int, max_row: int, min_col: int, max_col: int) -> dict:
    """Returns the formulas, cell reference, and 1-based row/column
    indices for all formulas in the sheet.

    If no formulas found in the sheet, returns None.

    Note: Must have a sheet from open_wkbk where formula_values = False

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to find formulas.

    min_row_data : int
        First row not containing headers.
    
    max_row : int
        Last row containing data.

    min_col : int
        First column containing data.

    max_col : int
        Last column containing data.

    Returns
    -------
    dict -> {<cell_ref: str>: [(<row_index: int>, <col_index: int>), <formula: str>]}
        Returns a dictionary containing the cell reference ('A1' for
        example) as the keys paired with lists containing the row and
        column indices (1-based) as well as the formula for each
        formula found in the sheet.
    """
    data_raw = sheet.iter_rows(
        min_row=min_row_data,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col
    )
    formulas = {}
    for row_index, row in enumerate(data_raw):
        for col_index, cell in enumerate(row):
            if type(cell.value) == type(''):
                if cell.value[0] == '=':
                    cell_ref = get_column_letter(col_index + min_col)
                    cell_ref += str(row_index + min_row_data)
                    formulas[cell_ref] = [row_index + min_row_data]
                    formulas[cell_ref].append(col_index + min_col)
                    formulas[cell_ref].append(cell.value)
    if len(formulas.keys()) > 0:
        return formulas
    return None

def sheet_check_formula_error(sheet: Worksheet, row_index: int, col_index: int) -> bool:
    """Returns boolean based on whether the formula at a given cell in
    the sheet returns an error.

    Note: Must have a sheet from open_wkbk where formula_values = True

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to check for an error returned by a formula.

    row_index : int
        1-based row index of a cell containing a formula.

    col_index : int
        1-based column index of a cell containing a formula.

    Returns
    -------
    bool
        If the cell contains an error returns True, otherwise returns
        False.
    """
    value = sheet.cell(row=row_index, column=col_index).value
    if type(value) == type(''):
        if value[0] in ['#']:
            return True
    return False

def sheet_hierarchical_fields(sheet: Worksheet, min_row: int, header_row: int, min_col: int, max_col: int) -> list:
    """Returns a list of hierarchical fields, compounded with '.'
    between each layer.

    If there is only one layer of fields (i.e. not hierarchical), it
    will return the only layer of fields.

    Parameters
    ----------
    sheet : Worksheet
        Worksheet containing hierarchical fields.

    min_row : int
        1-based row index for the first non-empty row in the sheet.

    header_row : int
        1-based row index for the last non-data row in the sheet.

    min_col : int
        1-based column index for the first non-empty column in the
        sheet.

    max_col : int
        1-based column indes for the last non-empty column in the
        sheet.

    Returns
    -------
    list -> [<layer1.layer2.->.layerN: str>]
        Returns a list of compounded hierarchical fields, separated by
        a '.' for each layer.
    """
    # Collects all fields from first non-empty row to the last row 
    # before data starts (header_row)
    headers = []
    for row_index in range(min_row, header_row + 1):
        headers.append(sheet_field_headers(sheet, row_index, min_col, max_col))
    
    # Populates None header with previous header for all but bottom row of headers
    header_temp = ''
    for r_index, row in enumerate(headers):
        if r_index < len(headers) - 1:
            for h_index, header in enumerate(row):
                if header is not None:
                    header_temp = header
                elif len(header_temp) > 0:
                    headers[r_index][h_index] = header_temp
    
    # Adds all layers from second last to top to bottom layer of headers, separated by '.'
    hierarchical_fields = headers[-1]
    for r_index, row in enumerate(reversed(headers)):
        if r_index > 0:
            for h_index, header in enumerate(row):
                if header is not None:
                    hierarchical_fields[h_index] = header + '.' + hierarchical_fields[h_index]

    return hierarchical_fields
